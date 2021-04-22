# coding=utf-8
import os
from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog
from PyQt5.uic import loadUi
import openpyxl as xl
from threading import Thread
import datetime

# 设置全局变量
filePath1 = ''
filePath2 = ''
filePath3 = ''
flag = False
class X2():
    global wb_result
    dir = os.path.abspath('..')
    wb_result = xl.Workbook(r"{}/data/核查结果.xlsx".format(dir))
    def __init__(self):
        dir = os.path.abspath('..')
        self.ui = loadUi(r"{}/UI/ZX_X2.ui".format(dir))
        # 加载lable
        self.ui.setWindowTitle("X2链路核查")

        self.ui.pushButton_noX2.clicked.connect(self.noX2)
        self.ui.pushButton_difPCI.clicked.connect(self.difPCI)
        self.ui.pushButton_reducNB.clicked.connect(self.reducNC)
        self.ui.pushButton_reducEB.clicked.connect(self.reducEC)
        self.ui.pushButton_takeALL.clicked.connect(self.takeALL)
        self.ui.pushButton_saveData.clicked.connect(self.saveData)

        # 打开文件按钮
        self.ui.pushButton_openNR.clicked.connect(self.openNR)
        self.ui.pushButton_openEB.clicked.connect(self.openEB)
        self.ui.pushButton_openLink.clicked.connect(self.openLink)

    # 打开NR台账
    def openNR(self):
        global filePath1
        filePath1, _ = QFileDialog.getOpenFileName(
            self.ui,
            "选择文件为",
            r"E:\03DWTool\X2\X2\data",
            "*.xlsx"
        )
        # filePath1 = r'data/无线中心nr台账（中兴+华为）20201217.xlsx'
        self.ui.lineEdit_openNR.setText(filePath1)
        # return filePath1
    # 打开外部小区文件
    def openEB(self):
        global filePath2
        filePath2, _ = QFileDialog.getOpenFileName(
            self.ui,
            "选择文件为",
            r"E:\03DWTool\X2\X2\data",
            "*.xlsx"
        )
        # filePath2 = r'data/SDR_fdd_radio.xlsx'
        self.ui.lineEdit_openEB.setText(filePath2)
        return filePath2
    # 打开链路文件
    def openLink(self):
        global filePath3
        filePath3, _ = QFileDialog.getOpenFileName(
            self.ui,
            "选择文件为",
            r"E:\03DWTool\X2\X2\data",
            "*.xlsx"
        )
        # filePath3 = r'data/SDR_sdr3.xlsx'
        self.ui.lineEdit_openLink.setText(filePath3)
        return filePath3
    # 检查是否全部打开需要文件
    def check_file(self):
        global flag
        if filePath1.strip() == '':
            flag = False
        elif filePath2.strip() == '':
            flag = False
        elif filePath3.strip() == '':
            flag = False
        else:
            flag = True
        if flag == False:
            QMessageBox.critical(
                self.ui,
                "错误",
                "没有打开文件"
            )
        else:
            self.ui.textBrowser.append('文件打开，可以计算')

    # 操作函数
    def noX2(self):
        X2.check_file(self)
        def noX2_check():
            self.ui.textBrowser.append('有邻区无链路计算中......')
            i = 0
            # 新建文档
            ws_noX2 = wb_result.create_sheet('有邻区无链路')
            result_list = []

            # 打开台账
            wb_stand = xl.load_workbook(filePath1)
            ws_standB = wb_stand['NR基站']
            standB_ip = ws_standB['E']  # 台账ip
            standB_gnodB1 = ws_standB['F']  # 台账基站
            standB_gnodB = []       # 基站索引
            for standB_gnodB1index in standB_gnodB1:
                standB_gnodB.append(str(standB_gnodB1index.value))

            # 打开外部关系
            wb_EB = xl.load_workbook(filePath2)
            ws_EB = wb_EB['ExternalNrCell']
            EB_enodB = ws_EB['E']  # 外部4g主小区
            EB_gnodB = ws_EB['J']  # 外部5g目标基站
            EB_gnodeBIP = []       # 链路索引

            # 打开链路文件
            wb_Link = xl.load_workbook(filePath3)
            ws_link = wb_Link['Sctp']
            Link_gnodB = ws_link['D']
            Link_IP = ws_link['O']
            Link_type = ws_link['AD']
            Link_gnodBIP = []
            self.ui.textBrowser.append('读取数据完成')
            # 外部关系链路索引
            for EB_enodB1 in EB_enodB:
                if str(EB_gnodB[EB_enodB1.row - 1].value) in standB_gnodB:
                    EB_gnodeBIP.append(str(EB_enodB1.value) + '_' +
                                       str(standB_ip[standB_gnodB.index(str(EB_gnodB[EB_enodB1.row - 1].value))].value))
            # 链路索引
            for Link_gnodBindex in Link_gnodB:
                if str(Link_type[Link_gnodBindex.row-1].value.strip()) == '2':
                    Link_gnodBIP.append(str(Link_gnodBindex.value) + '_' + str(Link_IP[Link_gnodBindex.row - 1].value))

            # 数据处理
            for EB_gnodeBIPindex in EB_gnodeBIP:
                if EB_gnodeBIPindex not in Link_gnodBIP:
                    result_list.append(EB_gnodeBIPindex)
            result_list = list(set(result_list))
            for result_listindex in result_list:
                result = result_listindex.split('_')
                ws_noX2.append([result[0], result[1]])

            i = len(result_list)
            self.ui.textBrowser.append(f'无X2有{i}项')
            flag = False
        if flag == True:
            thread = Thread(target=noX2_check)
            thread.start()


    def difPCI(self):
        X2.check_file(self)
        def difPCI_check():
            self.ui.textBrowser.append('PCI不一致计算中...')
            i = 0
            # 新建文档
            ws_difPCI = wb_result.create_sheet('外部PCI不一致')

            # 打开台账
            wb_stand = xl.load_workbook(filePath1)
            ws_standC = wb_stand['NR小区']
            stand_ECID1 = ws_standC['F']  # 台账小区
            stand_ECID = []
            for stand_ECID1ind in stand_ECID1:
                stand_ECID.append(stand_ECID1ind.value)
            stand_PCI = ws_standC['N']  # 台账小区PCI
            self.ui.textBrowser.append("读取台账成功")

            # 打开外部关系
            wb_EB = xl.load_workbook(filePath2)
            ws_EB = wb_EB['ExternalNrCell']
            EB_enodB = ws_EB['E']  # 外部4g主小区
            EB_gnodB = ws_EB['J']  # 外部5g目标基站
            EB_gCell = ws_EB['L']  # 外部5g目标基站小区
            EB_PCI = ws_EB['AE']  # 外部5g目标基站小区PCI
            # ws_difPCI.append(['4g主小区', '5g目标小区', '目标小区PCI', "台账PCI"])
            ws_difPCI.append(ws_EB[1])
            ws_difPCI.append(ws_EB[2])
            ws_difPCI.append(ws_EB[3])
            ws_difPCI.append(ws_EB[4])
            ws_difPCI.append(ws_EB[5])
            EB_gcID = []
            for EB_gindex in EB_gnodB:
                EB_gcID.append(str(EB_gindex.value) + '_' + str(EB_gCell[EB_gindex.row - 1].value))
            # 核查pci一致性
            for EB_enodBindex in EB_enodB:
                if EB_gcID[EB_enodBindex.row - 1] in stand_ECID:
                    if str(EB_PCI[EB_enodBindex.row - 1].value) != str(stand_PCI[stand_ECID.index(EB_gcID[EB_enodBindex.row - 1])].value):
                        ws_difPCI.append(ws_EB[EB_enodBindex.row - 1])
                        # ws_difPCI.append([EB_enodBindex.value, EB_gcID[EB_enodBindex.row - 1],
                        #                   EB_PCI[EB_enodBindex.row - 1].value,
                        #                   stand_PCI[stand_ECID.index(EB_gcID[EB_enodBindex.row - 1])].value
                        #                   ])
                        print(EB_PCI[EB_enodBindex.row - 1].value, '\t', stand_PCI[stand_ECID.index(EB_gcID[EB_enodBindex.row - 1])].value)
                        i = i + 1
            self.ui.textBrowser.append(f'PCI核查为{i}项')
            wb_result.close()
            flag = False
        if flag == True:
            thread = Thread(target=difPCI_check)
            thread.start()

    def reducNC(self):
        X2.check_file(self)
        def reducNC_check():
            self.ui.textBrowser.append('冗余邻区计算中......')
            i = 0
            # 新建文档
            ws_reducEB = wb_result.create_sheet('冗余邻区')

            # 打开台账
            wb_stand = xl.load_workbook(filePath1)
            ws_standC = wb_stand['NR小区']
            stand_ECID1 = ws_standC['F']  # 台账小区
            stand_ECID = []
            for ws_standC1index in stand_ECID1:
                stand_ECID.append(str(ws_standC1index.value))

            # 打开外部关系
            wb_EB = xl.load_workbook(filePath2)
            ws_EB = wb_EB['ExternalNrCell']
            EB_enodB = ws_EB['E']  # 外部4g主小区
            EB_gnodB1 = ws_EB['J']  # 外部5g目标基站
            EB_gCell = ws_EB['L']  # 外部5g目标基站小区
            EB_gnodB = []
            for EB_gnodBindex in EB_gnodB1:
                EB_gnodB.append(str(EB_gnodBindex.value) + "_" + str(EB_gCell[EB_gnodBindex.row - 1].value))
            for EB_enodBindex in EB_enodB:
                if EB_gnodB[EB_enodBindex.row - 1] not in stand_ECID:
                    ws_reducEB.append(ws_EB[EB_enodBindex.row])
                    i = i + 1
            self.ui.textBrowser.append(f"冗余邻区核查{i}项目")
            flag = False
        if flag == True:
            thread = Thread(target=reducNC_check)
            thread.start()

    def reducEC(self):
        X2.check_file(self)
        def reducEC_check():
            self.ui.textBrowser.append('冗余外部小区')
            i = 0
            # 新建文档
            # wb_result = xl.Workbook('./data/核查结果.xlsx')
            ws_reducEB = wb_result.create_sheet('冗余的外部小区')

            # 打开台账
            wb_stand = xl.load_workbook(filePath1)
            ws_standB = wb_stand['NR基站']
            standB1_gnodB = ws_standB['F']  # 台账基站
            standB_gnodB = []
            for standB_gnodBindex in standB1_gnodB:
                standB_gnodB.append(str(standB_gnodBindex.value))

            # 打开外部关系
            wb_EB = xl.load_workbook(filePath2)
            ws_EB = wb_EB['ExternalNrCell']
            EB_enodB = ws_EB['E']  # 外部4g主小区
            EB_gnodB = ws_EB['J']  # 外部5g目标基站
            EB_gCell = ws_EB['L']  # 外部5g目标基站小区
            for EB_gnodBindex in EB_gnodB:
                if EB_gnodBindex.value not in standB_gnodB:
                    ws_reducEB.append(ws_EB[EB_gnodBindex.row])
                    i = i + 1
            self.ui.textBrowser.append(f'核查冗余外部小区{i}')
            wb_result.close()
            flag = False
        if flag == True:
            thread = Thread(target=reducEC_check)
            thread.start()



    def takeALL(self):
        X2.check_file()
        if flag == True:
            self.ui.textBrowser.append('全部执行中......')
            X2.noX2()
            X2.reducEC()
            X2.difPCI()
            X2.reducNC()
        self.ui.textBrowser.append("全部计算完成！")
    def get_time(time = datetime.datetime.now()):
        print(time)
        new_time = str(time)
        now_time = new_time[1:19]
        return "".join(now_time)


    def saveData(self):
        dir = os.path.abspath('..')
        wb_result.save(f'{dir}\data\核查结果.xlsx')
        wb_result.close()
        self.ui.textBrowser.append('*********保留数据成功************')

